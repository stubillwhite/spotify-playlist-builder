from openpyxl.utils import get_column_letter, column_index_from_string

import openpyxl
import pandas as pd
import re

def column_index(address):
    match = re.match(r'([A-Z]+)(\d+)', address)
    return column_index_from_string(match[1])

def row_index(address):
    match = re.match(r'([A-Z]+)(\d+)', address)
    return int(match[2])

def partition(coll, n):
    for i in range(0, len(coll), n):
        yield coll[i:i+n]

def cells_from_range(cell_range):
    (start, end) = cell_range.split(':')
    cell_coordinates = [get_column_letter(col) + str(row)
            for row in range(row_index(start), row_index(end) + 1)
            for col in range(column_index(start), column_index(end) + 1)]

    row_length = column_index(end) + 1 - column_index(start) 
    return partition(cell_coordinates, row_length)

def values_from_range(sheet, cell_range):
    cell_coordinates = list(cells_from_range(cell_range))
    row_length = len(cell_coordinates[0])
    cell_values = [sheet[cell].value for row in cell_coordinates for cell in row]
    return list(partition(cell_values, row_length))

########################################
# Main
########################################

ANSWERS  = 'B3:E56'
GUESSES  = 'F3:Q56'
LIKES    = 'S3:AD56'
DISLIKES = 'AF3:AQ56'

if __name__=='__main__':

    dataframe = openpyxl.load_workbook('recs-roadtrip-playlist-quiz-master.xlsx')
    sheet = dataframe.active

    answers = values_from_range(sheet, ANSWERS)
    guesses = values_from_range(sheet, GUESSES)
    likes = values_from_range(sheet, LIKES)
    dislikes = values_from_range(sheet, DISLIKES)

    urls = [row[0] for row in answers[1:]]
    guessers = guesses[0]

    header = ['url', 'artist', 'track', 'submitter', 'guesser', 'guess', 'like', 'dislike']
    tbl = []
    for url in urls:
        for guesser in guessers:
            row_idx   = urls.index(url)+1
            artist    = answers[row_idx][1]
            track     = answers[row_idx][2]
            submitter = answers[row_idx][3]
            guess     = guesses[row_idx][guessers.index(guesser)]
            like      = likes[row_idx][guessers.index(guesser)]
            dislike   = dislikes[row_idx][guessers.index(guesser)]

            row = [url, artist, track, submitter, guesser, guess, like, dislike]
            tbl.append(row)

    df = pd.DataFrame(tbl, columns = header)
    print(df.to_string())
    df.to_csv('results.csv', index=False)


